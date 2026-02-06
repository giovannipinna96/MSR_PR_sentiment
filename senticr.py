"""
SentiCR - Sentiment Analysis for Code Review Comments

This module implements SentiCR (Ahmed et al., 2017), a sentiment analysis tool
specifically designed for code review interactions.

Reference:
    Ahmed, T., Bosu, A., Iqbal, A., & Rahimi, S. (2017). SentiCR: A customized
    sentiment analysis tool for code review interactions. In 2017 32nd IEEE/ACM
    International Conference on Automated Software Engineering (ASE) (pp. 106-111).

Original repository: https://github.com/senticr/SentiCR

This implementation:
- Downloads required data files automatically
- Provides a scikit-learn compatible interface
- Caches trained models for faster subsequent runs
- Offers batch prediction capabilities
"""

import os
import re
import pickle
import urllib.request
from pathlib import Path
from typing import List, Tuple, Optional, Union
import warnings

import numpy as np
import nltk
from nltk.stem.snowball import SnowballStemmer

# Ensure NLTK data is available
try:
    nltk.data.find('tokenizers/punkt_tab')
except LookupError:
    nltk.download('punkt_tab', quiet=True)

try:
    nltk.data.find('tokenizers/punkt')
except LookupError:
    nltk.download('punkt', quiet=True)

# ML imports
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.ensemble import GradientBoostingClassifier, RandomForestClassifier, AdaBoostClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.naive_bayes import MultinomialNB
from sklearn.linear_model import SGDClassifier
from sklearn.svm import LinearSVC
from sklearn.neural_network import MLPClassifier

# For handling imbalanced data
try:
    from imblearn.over_sampling import SMOTE
    SMOTE_AVAILABLE = True
except ImportError:
    SMOTE_AVAILABLE = False
    warnings.warn("imblearn not available. SMOTE will be disabled.")

# For reading Excel files
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    try:
        import xlrd
        XLRD_AVAILABLE = True
    except ImportError:
        XLRD_AVAILABLE = False


class SentiCRData:
    """Container for sentiment data."""
    def __init__(self, text: str, rating: int):
        self.text = text
        self.rating = rating  # -1 = negative, 0 = neutral, 1 = positive


class SentiCR:
    """
    SentiCR: Sentiment Analysis Tool for Code Review Comments.

    This classifier is specifically trained on code review data and handles
    code-specific patterns like URLs, code snippets, and technical jargon.

    Parameters
    ----------
    algo : str, default="GBT"
        Algorithm to use. Options: "GBT", "RF", "ADB", "DT", "NB", "SGD", "SVC", "MLPC"
    data_dir : str, optional
        Directory to store/load data files. Default: same directory as this file.
    use_smote : bool, default=True
        Whether to use SMOTE for handling class imbalance.
    cache_model : bool, default=True
        Whether to cache the trained model for faster loading.

    Attributes
    ----------
    model : sklearn classifier
        The trained classifier.
    vectorizer : TfidfVectorizer
        The fitted TF-IDF vectorizer.
    """

    # URLs for downloading required files
    DATA_URLS = {
        'oracle.xlsx': 'https://raw.githubusercontent.com/senticr/SentiCR/master/SentiCR/oracle.xlsx',
        'Contractions.txt': 'https://raw.githubusercontent.com/senticr/SentiCR/master/SentiCR/Contractions.txt',
        'EmoticonLookupTable.txt': 'https://raw.githubusercontent.com/senticr/SentiCR/master/SentiCR/EmoticonLookupTable.txt'
    }

    # Stop words list (from original SentiCR)
    STOP_WORDS = [
        'i', 'me', 'my', 'myself', 'we', 'our', 'ours', 'ourselves', 'you', 'your',
        'yours', 'yourself', 'yourselves', 'he', 'him', 'his', 'himself', 'she', 'her',
        'hers', 'herself', 'it', 'its', 'itself', 'they', 'them', 'their', 'theirs',
        'themselves', 'what', 'which', 'who', 'whom', 'this', 'that', 'these', 'those',
        'am', 'is', 'are', 'was', 'were', 'be', 'been', 'being', 'have', 'has', 'had',
        'having', 'do', 'does', 'did', 'doing', 'a', 'an', 'the', 'and', 'but', 'if',
        'or', 'because', 'as', 'until', 'while', 'of', 'at', 'by', 'for', 'with',
        'about', 'against', 'between', 'into', 'through', 'during', 'before', 'after',
        'above', 'below', 'to', 'from', 'up', 'down', 'in', 'out', 'on', 'off', 'over',
        'under', 'again', 'further', 'then', 'once', 'here', 'there', 'when', 'where',
        'why', 'how', 'all', 'each', 'few', 'more', 'most', 'other', 'some', 'such',
        'only', 'own', 'same', 'so', 'than', 'too', 'very', 's', 't', 'can', 'will',
        'just', 'don', 'should', 'now', 'd', 'll', 'm', 'o', 're', 've', 'y', 'ain',
        'aren', 'couldn', 'didn', 'doesn', 'hadn', 'hasn', 'haven', 'isn', 'ma',
        'mightn', 'mustn', 'needn', 'shan', 'shouldn', 'wasn', 'weren', 'won', 'wouldn',
        # Programming-related stop words
        'code', 'file', 'line', 'function', 'method', 'class', 'variable', 'return',
        'null', 'true', 'false', 'int', 'string', 'void', 'public', 'private', 'static'
    ]

    # Available algorithms
    ALGORITHMS = {
        'GBT': GradientBoostingClassifier,
        'RF': RandomForestClassifier,
        'ADB': AdaBoostClassifier,
        'DT': DecisionTreeClassifier,
        'NB': MultinomialNB,
        'SGD': SGDClassifier,
        'SVC': LinearSVC,
        'MLPC': MLPClassifier
    }

    def __init__(
        self,
        algo: str = "GBT",
        data_dir: Optional[str] = None,
        use_smote: bool = True,
        cache_model: bool = True
    ):
        self.algo = algo
        self.use_smote = use_smote and SMOTE_AVAILABLE
        self.cache_model = cache_model

        # Set data directory
        if data_dir is None:
            self.data_dir = Path(__file__).parent / "senticr_data"
        else:
            self.data_dir = Path(data_dir)
        self.data_dir.mkdir(parents=True, exist_ok=True)

        # Initialize components
        self.model = None
        self.vectorizer = None
        self.contractions = {}
        self.emoticons = {}
        self.stemmer = SnowballStemmer("english")

        # Load or download data files
        self._ensure_data_files()
        self._load_lookup_tables()

        # Load cached model or train new one
        self._load_or_train_model()

    def _ensure_data_files(self):
        """Download required data files if not present."""
        for filename, url in self.DATA_URLS.items():
            filepath = self.data_dir / filename
            if not filepath.exists():
                print(f"   Downloading {filename}...")
                try:
                    urllib.request.urlretrieve(url, filepath)
                except Exception as e:
                    raise RuntimeError(f"Failed to download {filename}: {e}")

    def _load_lookup_tables(self):
        """Load contractions and emoticon lookup tables."""
        # Load contractions
        contractions_path = self.data_dir / "Contractions.txt"
        if contractions_path.exists():
            with open(contractions_path, 'r', encoding='utf-8', errors='ignore') as f:
                for line in f:
                    parts = line.strip().split(':')
                    if len(parts) == 2:
                        self.contractions[parts[0].strip()] = parts[1].strip()

        # Load emoticons
        emoticons_path = self.data_dir / "EmoticonLookupTable.txt"
        if emoticons_path.exists():
            with open(emoticons_path, 'r', encoding='utf-8', errors='ignore') as f:
                for line in f:
                    parts = line.strip().split('\t')
                    if len(parts) == 2:
                        self.emoticons[parts[0].strip()] = parts[1].strip()

    def _load_or_train_model(self):
        """Load cached model or train a new one."""
        cache_path = self.data_dir / f"senticr_model_{self.algo}.pkl"

        if self.cache_model and cache_path.exists():
            print(f"   Loading cached SentiCR model ({self.algo})...")
            with open(cache_path, 'rb') as f:
                cached = pickle.load(f)
                self.model = cached['model']
                self.vectorizer = cached['vectorizer']
            return

        # Train new model
        print(f"   Training SentiCR model ({self.algo})...")
        training_data = self._read_oracle_data()
        self._train_model(training_data)

        # Cache model
        if self.cache_model:
            with open(cache_path, 'wb') as f:
                pickle.dump({
                    'model': self.model,
                    'vectorizer': self.vectorizer
                }, f)
            print(f"   Model cached at: {cache_path}")

    def _read_oracle_data(self) -> List[SentiCRData]:
        """Read training data from oracle.xlsx."""
        oracle_path = self.data_dir / "oracle.xlsx"
        data = []

        if OPENPYXL_AVAILABLE:
            wb = load_workbook(oracle_path, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
                if row[0] and row[1] is not None:
                    text = str(row[0])
                    rating = int(row[1])
                    data.append(SentiCRData(text, rating))
            wb.close()
        elif XLRD_AVAILABLE:
            import xlrd
            wb = xlrd.open_workbook(oracle_path)
            ws = wb.sheet_by_index(0)
            for i in range(1, ws.nrows):  # Skip header
                text = str(ws.cell_value(i, 0))
                rating = int(ws.cell_value(i, 1))
                data.append(SentiCRData(text, rating))
        else:
            raise RuntimeError("Neither openpyxl nor xlrd is available for reading Excel files")

        print(f"   Loaded {len(data)} training samples from oracle")
        return data

    def _preprocess_text(self, text: str) -> str:
        """
        Preprocess text for sentiment analysis.

        Steps:
        1. Encode to ASCII (remove special chars)
        2. Expand contractions
        3. Remove URLs
        4. Replace emoticons with sentiment words
        5. Handle negation
        """
        if not isinstance(text, str):
            text = str(text)

        # Encode to ASCII
        text = text.encode('ascii', 'ignore').decode('ascii')

        # Expand contractions
        for contraction, expansion in self.contractions.items():
            text = text.replace(contraction, expansion)

        # Remove URLs
        text = re.sub(r'http[s]?://\S+', ' ', text)
        text = re.sub(r'www\.\S+', ' ', text)

        # Replace emoticons
        for emoticon, sentiment in self.emoticons.items():
            text = text.replace(emoticon, f' {sentiment} ')

        # Handle negation (attach NOT_ prefix to following words)
        negation_words = ['not', 'no', 'never', 'neither', 'nobody', 'nothing',
                         'nowhere', 'none', "n't", 'cannot', 'cant', 'wont', 'dont']
        words = text.lower().split()
        result = []
        negate = False

        for word in words:
            if word in negation_words:
                negate = True
                result.append(word)
            elif negate:
                if word in ['.', '!', '?', ',', ';', ':']:
                    negate = False
                    result.append(word)
                else:
                    result.append(f'NOT_{word}')
            else:
                result.append(word)

        return ' '.join(result)

    def _tokenize_and_stem(self, text: str) -> str:
        """Tokenize, stem, and rejoin text for TF-IDF."""
        tokens = nltk.word_tokenize(text.lower())
        stemmed = [self.stemmer.stem(token) for token in tokens if token.isalpha()]
        return ' '.join(stemmed)

    def _train_model(self, training_data: List[SentiCRData]):
        """Train the sentiment model."""
        # Preprocess all texts and apply stemming
        texts = [self._tokenize_and_stem(self._preprocess_text(d.text)) for d in training_data]
        labels = [d.rating for d in training_data]

        # Create TF-IDF vectorizer (without custom tokenizer for pickle compatibility)
        # Note: min_df=1 to include rare but important sentiment words
        # The original SentiCR used min_df=3 but with only 1599 samples this
        # removes too many important terms
        self.vectorizer = TfidfVectorizer(
            stop_words=self.STOP_WORDS,
            max_df=0.5,
            min_df=1,  # Changed from 3 to preserve rare sentiment words
            sublinear_tf=True,
            use_idf=True
        )

        # Transform texts to features
        X = self.vectorizer.fit_transform(texts).toarray()
        y = np.array(labels)

        # Apply SMOTE for class balancing
        if self.use_smote and len(set(y)) > 1:
            try:
                smote = SMOTE(sampling_strategy=0.5, k_neighbors=min(5, min(np.bincount(y + 1)) - 1))
                X, y = smote.fit_resample(X, y)
                print(f"   Applied SMOTE: {len(y)} samples after resampling")
            except Exception as e:
                print(f"   SMOTE failed, using original data: {e}")

        # Get classifier
        classifier_class = self.ALGORITHMS.get(self.algo, GradientBoostingClassifier)

        # Set algorithm-specific parameters
        if self.algo == 'GBT':
            self.model = classifier_class(n_estimators=100, random_state=42)
        elif self.algo == 'RF':
            self.model = classifier_class(n_estimators=100, random_state=42)
        elif self.algo == 'MLPC':
            self.model = classifier_class(hidden_layer_sizes=(100,), max_iter=500, random_state=42)
        elif self.algo == 'SVC':
            self.model = classifier_class(max_iter=1000, random_state=42)
        else:
            self.model = classifier_class()

        # Train
        self.model.fit(X, y)
        print(f"   Model trained successfully")

    def get_sentiment_polarity(self, text: str) -> int:
        """
        Get sentiment polarity for a single text.

        Parameters
        ----------
        text : str
            The text to analyze.

        Returns
        -------
        int
            Sentiment class: -1 (negative), 0 (neutral), 1 (positive)
        """
        processed = self._tokenize_and_stem(self._preprocess_text(text))
        features = self.vectorizer.transform([processed]).toarray()
        return int(self.model.predict(features)[0])

    def get_sentiment_polarity_collection(self, texts: List[str]) -> List[int]:
        """
        Get sentiment polarity for multiple texts.

        Parameters
        ----------
        texts : List[str]
            List of texts to analyze.

        Returns
        -------
        List[int]
            List of sentiment classes.
        """
        processed = [self._tokenize_and_stem(self._preprocess_text(t)) for t in texts]
        features = self.vectorizer.transform(processed).toarray()
        return [int(p) for p in self.model.predict(features)]

    def predict_proba(self, texts: Union[str, List[str]]) -> np.ndarray:
        """
        Get probability scores for each class.

        Parameters
        ----------
        texts : str or List[str]
            Text(s) to analyze.

        Returns
        -------
        np.ndarray
            Probability scores of shape (n_samples, n_classes).
            Classes are ordered: [-1, 0, 1] (negative, neutral, positive)
        """
        if isinstance(texts, str):
            texts = [texts]

        processed = [self._tokenize_and_stem(self._preprocess_text(t)) for t in texts]
        features = self.vectorizer.transform(processed).toarray()

        if hasattr(self.model, 'predict_proba'):
            probs = self.model.predict_proba(features)
            return probs
        else:
            # For models without predict_proba, return one-hot based on prediction
            predictions = self.model.predict(features)
            n_classes = len(self.model.classes_)
            probs = np.zeros((len(predictions), n_classes))
            for i, pred in enumerate(predictions):
                class_idx = np.where(self.model.classes_ == pred)[0][0]
                probs[i, class_idx] = 1.0
            return probs

    def get_sentiment_label(self, polarity: int) -> str:
        """
        Convert polarity to label string.

        Note: SentiCR is trained as a binary classifier:
        - -1 = negative (clearly negative sentiment)
        - 0 = non-negative (neutral or positive, mapped to "neutral" for compatibility)

        There is no explicit "positive" class in the original SentiCR training data.
        """
        if polarity == -1:
            return "negative"
        else:
            return "neutral"  # Non-negative (includes neutral and positive)

    def analyze(self, texts: Union[str, List[str]]) -> List[dict]:
        """
        Analyze sentiment with detailed output (similar to HuggingFace pipeline).

        Parameters
        ----------
        texts : str or List[str]
            Text(s) to analyze.

        Returns
        -------
        List[dict]
            List of dicts with 'label' and 'score' keys.
        """
        if isinstance(texts, str):
            texts = [texts]

        polarities = self.get_sentiment_polarity_collection(texts)

        # Try to get probabilities
        try:
            probs = self.predict_proba(texts)
            has_probs = True
        except:
            has_probs = False

        results = []
        for i, polarity in enumerate(polarities):
            label = self.get_sentiment_label(polarity)
            if has_probs and hasattr(self.model, 'classes_'):
                class_idx = np.where(self.model.classes_ == polarity)[0][0]
                score = float(probs[i, class_idx])
            else:
                score = 1.0

            results.append({
                'label': label,
                'score': score,
                'polarity': polarity
            })

        return results


def create_senticr_pipeline(algo: str = "GBT", data_dir: Optional[str] = None):
    """
    Create a SentiCR instance configured like a HuggingFace pipeline.

    This function provides a convenient way to initialize SentiCR with
    default settings suitable for integration with other sentiment models.

    Parameters
    ----------
    algo : str, default="GBT"
        Algorithm to use.
    data_dir : str, optional
        Directory for data files.

    Returns
    -------
    SentiCR
        Initialized SentiCR instance.
    """
    return SentiCR(algo=algo, data_dir=data_dir, use_smote=True, cache_model=True)


# For testing
if __name__ == "__main__":
    print("Testing SentiCR...")

    # Initialize
    senticr = create_senticr_pipeline()

    # Test texts
    test_texts = [
        "This code looks great! Well done!",
        "This is a bug, you need to fix it immediately",
        "Please add documentation for this method",
        "I think this could cause a security vulnerability",
        "LGTM, nice refactoring!"
    ]

    print("\nTest Results:")
    print("-" * 60)

    for text in test_texts:
        result = senticr.analyze(text)[0]
        print(f"Text: {text[:50]}...")
        print(f"  -> {result['label']} (score: {result['score']:.3f})")
        print()
